import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

def split_name(full_name):
    parts = full_name.strip().split()
    for i in range(1, len(parts)):
        if parts[i][0].isupper() and not parts[i].isupper():
            surname = ' '.join(parts[:i])
            name = ' '.join(parts[i:])
            return surname.title(), surname.upper(), name.title()
    return parts[0].title(), parts[0].upper(), ' '.join(parts[1:]).title()

def process_excel_file(input_file, output_file):
    df = pd.read_excel(input_file)

    proper_surnames = []
    uppercase_surnames = []
    names = []

    for full_name in df.iloc[:, 2]:  # Column C
        if isinstance(full_name, str):
            surname, surname_upper, name = split_name(full_name)
        else:
            surname, surname_upper, name = "", "", ""
        proper_surnames.append(surname)
        uppercase_surnames.append(surname_upper)
        names.append(name)

    df.insert(3, 'Uppercase Surname', uppercase_surnames)
    df.iloc[:, 4] = proper_surnames  # Adjusted due to insertion
    df.iloc[:, 5] = names

    # Add a new column 'molteplicità' initialized empty
    df['molteplicità'] = ""

    # Group by full name
    full_name_col = df.iloc[:, 4] + " " + df.iloc[:, 5]  # Surname + Name
    df['__full_name__'] = full_name_col  # Temporary helper column

    # Process duplicates
    duplicated_names = df[df.duplicated('__full_name__', keep=False)]

    for name, group in duplicated_names.groupby('__full_name__'):
        ateneos = group.iloc[:, 10].unique()  # Column K = index 10
        indices = group.index

        if len(ateneos) > 1:
            df.loc[indices, 'molteplicità'] = 'diverso ateneo'
        else:
            df.loc[indices, 'molteplicità'] = 'omonimia'

    # Save intermediate result to Excel (no style yet)
    temp_file = "__temp_unstyled_output.xlsx"
    df.drop(columns='__full_name__').to_excel(temp_file, index=False)

    # Now re-open with openpyxl to apply coloring
    wb = load_workbook(temp_file)
    ws = wb.active

    cyan_fill = PatternFill(start_color="00FFFF", end_color="00FFFF", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

    molteplicità_col_index = df.columns.get_loc("molteplicità") + 1  # Excel is 1-indexed

    for row in range(2, ws.max_row + 1):
        value = ws.cell(row=row, column=molteplicità_col_index).value
        if value == "diverso ateneo":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = cyan_fill
        elif value == "omonimia":
            for col in range(1, ws.max_column + 1):
                ws.cell(row=row, column=col).fill = yellow_fill

    wb.save(output_file)

# Example usage
input_file = "INFO-01.xlsx"
output_file = "INFO-01_updated.xlsx"
process_excel_file(input_file, output_file)
